import streamlit as st
from pathlib import Path
import base64

# Initial page config

st.set_page_config(
     page_title='Streamlit cheat sheet',
     layout="wide",
     initial_sidebar_state="expanded",
)

def main():
    cs_sidebar()

    return None

# Thanks to streamlitopedia for the following code snippet

def img_to_bytes(img_path):
    img_bytes = Path(img_path).read_bytes()
    encoded = base64.b64encode(img_bytes).decode()
    return encoded

# sidebar

def cs_sidebar():

    st.sidebar.markdown('''[<img src='data:image/png;base64,{}' class='img-fluid' width=170 height=32>](https://streamlit.io/)'''.format(img_to_bytes("dataroots-logo.png")), unsafe_allow_html=True)
    st.sidebar.header('Career path')

    st.sidebar.markdown(''' Explore career path insights here ''', unsafe_allow_html=True)

    st.sidebar.markdown('__Line1__')

    st.sidebar.code('text 1')

    st.sidebar.markdown('__Instructions to use the app__')
    st.sidebar.markdown('You should start by copying your levels on different career path topics')

    st.sidebar.markdown('__Info 1__')


    st.sidebar.markdown('__Info 2__')
    st.sidebar.markdown('<small>Learn more about [career path levels](https://docs.streamlit.io/library/advanced-features/prerelease#beta-and-experimental-features)</small>', unsafe_allow_html=True)

    st.sidebar.markdown('''<hr>''', unsafe_allow_html=True)
    st.sidebar.markdown('''<small>[Career path sheet](https://github.com/daniellewisDL/streamlit-cheat-sheet)  | June 2024 | [Dataroots](https://dataroots.io/)</small>''', unsafe_allow_html=True)

    return None

    




## APPLICATION FOR INPUT/OUTPUT 
#button first
#df second
#add visuals (heatmap+)
#focus on progression

import streamlit as st
import pandas as pd
import subprocess
from openpyxl import load_workbook

# Function to check installed packages
def check_installed_packages():
    installed_packages = subprocess.check_output(['pip', 'freeze']).decode('utf-8')
    return installed_packages

# Function to check if a column is fully completed (i.e., all cells are empty)
def is_column_completed(column):
    return column.isna().all()

# Function to get the level and non-empty cells for a group of columns
def get_level_and_values(columns):
    level = 0
    non_empty_cells = []
    for col_name, col_data in columns.items():
        if not is_column_completed(col_data):
            level += 1
            non_empty_cells.extend(col_data.dropna().tolist())
            break  # Stop at the first non-empty column
    return level, non_empty_cells

# Function to extract cell colors from the worksheet
def get_cell_colors(ws):
    cell_colors = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor:
                color = cell.fill.fgColor.rgb
                cell_colors[(cell.row, cell.column)] = color
            else:
                cell_colors[(cell.row, cell.column)] = None
    return cell_colors

# Function to style the DataFrame with extracted cell colors
def style_dataframe(df, cell_colors):
    def apply_color(row):
        row_colors = []
        for i, cell in enumerate(row):
            color = cell_colors.get((row.name + 2, i + 1))  # Adjust for header and 1-based index
            if color:
                row_colors.append(f"background-color: #{color[2:]}")
            else:
                row_colors.append("")
        return row_colors
    
    return df.style.apply(apply_color, axis=1)

st.title("Data Career Path Level Up")

# Display installed packages
st.write("Installed Packages:")
st.text(check_installed_packages())

# File uploader
file = st.file_uploader("Upload Your Personal Career Path Excel file", type=['xlsx'])

if file is not None:
    try:
        # Load Excel data from the specified sheet
        df = pd.read_excel(file, sheet_name='Sheet1', engine='openpyxl')

        # Load workbook and worksheet to extract cell colors
        wb = load_workbook(file, data_only=True)
        ws = wb['Sheet1']
        cell_colors = get_cell_colors(ws)

        # Select only the columns of interest
        ae_columns = df.filter(like='AE', axis=1)
        ds_columns = df.filter(like='DS', axis=1)
        at_columns = df.filter(like='AT', axis=1)

        # Get level and non-empty cells for AE, DS, and AT columns
        ae_level, ae_values = get_level_and_values(ae_columns)
        ds_level, ds_values = get_level_and_values(ds_columns)
        at_level, at_values = get_level_and_values(at_columns)

        # Display domain buttons
        st.subheader("Select Domain to Display:")
        selected_domain = st.radio("For which domain do you want to improve?", ['AE', 'DS', 'AT'])

        # Debug: Display the filtered columns
        if selected_domain == 'AE':
            st.write("AE Columns:")
            styled_ae = style_dataframe(ae_columns, cell_colors)
            st.dataframe(styled_ae.to_html(), unsafe_allow_html=True)
        elif selected_domain == 'DS':
            st.write("DS Columns:")
            styled_ds = style_dataframe(ds_columns, cell_colors)
            st.dataframe(styled_ds.to_html(), unsafe_allow_html=True)
        elif selected_domain == 'AT':
            st.write("AT Columns:")
            styled_at = style_dataframe(at_columns, cell_colors)
            st.dataframe(styled_at.to_html(), unsafe_allow_html=True)
    except Exception as e:
        st.error(f"An error occurred: {e}")

        # remove info over current and next level!!
        #suggestive

        # Display results based on the selected domain
#        if selected_domain == 'AE':
#            st.subheader("On which domains do you need to improve to level up for the AE track?")
#            st.write(f"My current level: {ae_level}")
#            st.write("The following areas you should try to improve to level up again!:")
#            for value in ae_values:
#                st.write(value)
#        elif selected_domain == 'DS':
#            st.subheader("On which domains do you need to improve to level up for the DS track")
#            st.write(f"My current level: {ds_level}")
#            st.write("The following areas you should try to improve to level up again:")
#           for value in ds_values:
#                st.write(value)
#        elif selected_domain == 'AT':
#            st.subheader("On which domains do you need to improve to level up for the AT track")
#            st.write(f"My current level: {at_level}")
#            st.write("The following areas you should try to improve to level up again:")
#            for value in at_values:
#                st.write(value)
#    except Exception as e:
#        st.error(f"An error occurred: {e}")
    #add dynamic or simply space between the two parts, frames?


# Run main()

